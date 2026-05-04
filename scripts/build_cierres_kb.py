from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SHEET_NAME = "Registro de alarmas y eventos"
STORE_PATTERN = re.compile(r"^\s*(?P<code>\d+)\s*-\s*(?P<name>.+?)\s*$")


class RawEvent(dict[str, Any]):
    pass


def normalize_store_label(value: Any) -> str:
    return " ".join(str(value).strip().split())


def split_store(value: Any) -> tuple[str | None, str, str]:
    label = normalize_store_label(value)
    match = STORE_PATTERN.match(label)

    if not match:
        return None, label, label

    store_code = match.group("code")
    store_name = match.group("name").strip()
    return store_code, store_name, f"{store_code} - {store_name}"


def iso_date(value: datetime) -> str:
    return value.strftime("%Y-%m-%d")


def iso_time(value: datetime) -> str:
    return value.strftime("%H:%M:%S")


def iso_datetime(value: datetime) -> str:
    return value.strftime("%Y-%m-%dT%H:%M:%S")


def load_events(source_path: Path, sheet_name: str) -> tuple[str, list[RawEvent]]:
    workbook = load_workbook(source_path, data_only=True, read_only=True)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = next(
            sheet
            for sheet in workbook.worksheets
            if sheet.max_row > 1 and sheet.max_column >= 2
        )

    events: list[RawEvent] = []

    for row_number, (store_value, activated_at, *_) in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if store_value is None or activated_at is None:
            continue

        if not isinstance(activated_at, datetime):
            continue

        store_code, store_name, store_label = split_store(store_value)
        activation_date = iso_date(activated_at)
        record_id = f"{store_label}|{activation_date}"
        events.append(
            {
                "recordId": record_id,
                "storeCode": store_code,
                "storeName": store_name,
                "storeLabel": store_label,
                "date": activation_date,
                "time": iso_time(activated_at),
                "activationDateTime": iso_datetime(activated_at),
                "sourceRowNumber": row_number,
            }
        )

    return worksheet.title, events


def build_knowledge_base(source_path: Path, sheet_name: str) -> dict[str, Any]:
    resolved_source = source_path.resolve()
    selected_sheet_name, events = load_events(resolved_source, sheet_name)

    grouped_events: dict[tuple[str, str], list[RawEvent]] = defaultdict(list)
    exact_duplicates = Counter(
        (event["storeLabel"], event["activationDateTime"]) for event in events
    )

    for event in events:
        grouped_events[(event["storeLabel"], event["date"])].append(event)

    records: list[dict[str, Any]] = []
    discarded_events: list[dict[str, Any]] = []
    per_store_counter = Counter(event["storeLabel"] for event in events)

    for (_, _), bucket in sorted(grouped_events.items()):
        sorted_bucket = sorted(
            bucket,
            key=lambda event: (event["activationDateTime"], event["sourceRowNumber"]),
        )
        kept_event = sorted_bucket[-1]
        dropped_events = sorted_bucket[:-1]
        kept_datetime = datetime.fromisoformat(kept_event["activationDateTime"])

        records.append(
            {
                "recordId": kept_event["recordId"],
                "storeCode": kept_event["storeCode"],
                "storeName": kept_event["storeName"],
                "storeLabel": kept_event["storeLabel"],
                "date": kept_event["date"],
                "closingTime": kept_event["time"],
                "closingDateTime": kept_event["activationDateTime"],
                "sourceRowNumber": kept_event["sourceRowNumber"],
                "rawEventCount": len(sorted_bucket),
                "discardedCount": len(dropped_events),
                "discardedTimes": [event["time"] for event in dropped_events],
                "exactDuplicateCount": max(
                    0,
                    exact_duplicates[
                        (kept_event["storeLabel"], kept_event["activationDateTime"])
                    ]
                    - 1,
                ),
            }
        )

        for dropped_event in dropped_events:
            dropped_datetime = datetime.fromisoformat(
                dropped_event["activationDateTime"]
            )
            discarded_events.append(
                {
                    "recordId": kept_event["recordId"],
                    "storeCode": dropped_event["storeCode"],
                    "storeName": dropped_event["storeName"],
                    "storeLabel": dropped_event["storeLabel"],
                    "date": dropped_event["date"],
                    "time": dropped_event["time"],
                    "activationDateTime": dropped_event["activationDateTime"],
                    "sourceRowNumber": dropped_event["sourceRowNumber"],
                    "keptClosingTime": kept_event["time"],
                    "keptClosingDateTime": kept_event["activationDateTime"],
                    "secondsFromKeptClosing": int(
                        (dropped_datetime - kept_datetime).total_seconds()
                    ),
                    "reason": "same_store_same_day_repeated_event",
                }
            )

    records.sort(key=lambda item: (item["date"], item["closingTime"], item["storeLabel"]))
    discarded_events.sort(
        key=lambda item: (item["date"], item["time"], item["storeLabel"])
    )

    unique_store_labels = sorted({event["storeLabel"] for event in events})
    dates = [event["date"] for event in events]
    summary_by_date = []

    for date in sorted({record["date"] for record in records}):
        day_records = [record for record in records if record["date"] == date]
        summary_by_date.append(
            {
                "date": date,
                "cleanClosings": len(day_records),
                "discardedEvents": sum(
                    int(record["discardedCount"]) for record in day_records
                ),
                "distinctStores": len(
                    {record["storeLabel"] for record in day_records}
                ),
            }
        )

    stores = []

    for store_label in unique_store_labels:
        store_code, store_name, _ = split_store(store_label)
        stores.append(
            {
                "storeLabel": store_label,
                "storeCode": store_code,
                "storeName": store_name,
                "rawEventCount": per_store_counter[store_label],
                "cleanClosingCount": sum(
                    1 for record in records if record["storeLabel"] == store_label
                ),
            }
        )

    return {
        "meta": {
            "generatedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "sourceWorkbook": str(resolved_source),
            "sourceSheet": selected_sheet_name,
            "cleaningRule": "keep_last_activation_per_store_per_calendar_day",
            "rawEvents": len(events),
            "exactDuplicateRows": sum(
                count - 1 for count in exact_duplicates.values() if count > 1
            ),
            "cleanClosings": len(records),
            "discardedEvents": len(discarded_events),
            "distinctStores": len(unique_store_labels),
            "dateRange": {
                "start": min(dates) if dates else None,
                "end": max(dates) if dates else None,
            },
            "supportedCategories": ["close"],
        },
        "records": records,
        "discardedEvents": discarded_events,
        "stores": stores,
        "summaryByDate": summary_by_date,
    }


def write_json(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build a clean knowledge base of store closings from a source workbook."
    )
    parser.add_argument("source", help="Path to the source workbook.")
    parser.add_argument("output", help="Path to the JSON output file.")
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help="Worksheet name to import. Defaults to the alarms/events sheet.",
    )
    return parser


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()
    payload = build_knowledge_base(Path(args.source), args.sheet)
    write_json(payload, Path(args.output))
    print(
        json.dumps(
            {
                "status": "ok",
                "output": str(Path(args.output).resolve()),
                "cleanClosings": payload["meta"]["cleanClosings"],
                "discardedEvents": payload["meta"]["discardedEvents"],
                "distinctStores": payload["meta"]["distinctStores"],
            }
        )
    )


if __name__ == "__main__":
    main()
